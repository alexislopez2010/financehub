#!/usr/bin/env python3
"""
Migrate transactions and bills from Family_Finance_Dashboard_2026.xlsx into Supabase.

Usage:
  pip install openpyxl supabase python-dotenv
  export SUPABASE_URL=https://xxxx.supabase.co
  export SUPABASE_SERVICE_ROLE_KEY=eyJ...          # service-role key (NOT the anon key)
  export HOUSEHOLD_ID=00000000-0000-0000-0000-000000000001
  python migrate_from_excel.py /path/to/Family_Finance_Dashboard_2026.xlsx

The service-role key bypasses RLS so the script can insert rows for the
household directly. Keep it secret — do not commit it or put it in the browser.
"""
import os
import sys
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from supabase import create_client

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
HOUSEHOLD_ID = os.environ.get("HOUSEHOLD_ID", "00000000-0000-0000-0000-000000000001")

def fingerprint(date, desc, amount, account):
    raw = f"{date}|{desc}|{amount}|{account}".lower()
    return hashlib.sha256(raw.encode()).hexdigest()[:16]

def iso_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else None

def migrate_transactions(ws, sb):
    rows = []
    # Header is row 3, data starts row 4
    for r in ws.iter_rows(min_row=4, values_only=True):
        _, date, desc, amount, ttype, cat, account, member, pay, imported, notes = (r + (None,)*11)[:11]
        if not date or not desc or amount is None:
            continue
        d = iso_date(date)
        rows.append({
            "household_id": HOUSEHOLD_ID,
            "date": d,
            "description": str(desc).strip(),
            "amount": float(amount),
            "type": (ttype or "Expense").strip(),
            "category": cat,
            "account": account,
            "member": member,
            "payment_method": pay,
            "notes": notes,
            "fingerprint": fingerprint(d, desc, amount, account),
        })
    print(f"Transactions to upload: {len(rows)}")
    # Insert in batches of 500
    for i in range(0, len(rows), 500):
        batch = rows[i:i+500]
        sb.table("transactions").insert(batch).execute()
        print(f"  inserted {i+len(batch)}/{len(rows)}")

def migrate_bills(ws, sb):
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        name, cat, account, due_day, freq, budget = r[:6]
        if not name:
            continue
        rows.append({
            "household_id": HOUSEHOLD_ID,
            "name": str(name).strip(),
            "category": cat,
            "account": account,
            "due_day": int(due_day) if due_day else None,
            "frequency": freq or "Monthly",
            "budget_amount": float(budget or 0),
            "is_active": True,
        })
    print(f"Bills to upload: {len(rows)}")
    if rows:
        sb.table("bills").insert(rows).execute()

def main():
    if len(sys.argv) < 2:
        print("Usage: python migrate_from_excel.py <xlsx path>")
        sys.exit(1)
    wb = load_workbook(sys.argv[1], data_only=True)
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # OPTIONAL: wipe existing rows for a clean re-import
    if "--fresh" in sys.argv:
        print("Wiping existing transactions & bills for household…")
        sb.table("transactions").delete().eq("household_id", HOUSEHOLD_ID).execute()
        sb.table("bills").delete().eq("household_id", HOUSEHOLD_ID).execute()

    if "Transactions" in wb.sheetnames:
        migrate_transactions(wb["Transactions"], sb)
    if "Bills" in wb.sheetnames:
        migrate_bills(wb["Bills"], sb)

    print("Done.")

if __name__ == "__main__":
    main()
